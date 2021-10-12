
import datetime
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from pandas import ExcelFile
from pathlib import Path
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import xlsxwriter
from UliPlot.XLSX import auto_adjust_xlsx_column_width
from openpyxl.utils import get_column_letter

data_folder = Path("C:/MSIT/Spring 2021/Info sec/Project/Information Security _ Privacy Material-20210322T044634Z-001/Information Security _ Privacy Material/")
# denote spreadsheets
excel_file_1 = data_folder / "ajb9b3_upd.xlsx"
result_file = "C:/MSIT/Spring 2021/Info sec/Project/Information Security _ Privacy Material-20210322T044634Z-001/Results_10s.xlsx"

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

# def calculate_week():


def result_file_data(start_week, end_week):
    wbook = load_workbook(result_file)
    sheet = wbook["Sheet1"]

    sheet.merge_cells('B1:C1')
    sheet['B1'] = 'UserA'
    sheet['B1'].alignment = Alignment(horizontal='center')
    sheet['B2'] = 'Time'
    sheet['C2'] = 'Octets/Duration'
    cnt = 3
    wk_cnt = 3
    usr = 1
    i = 0
    for wk in start_week:
        start_of_wk = wk
        end_of_wk = end_week[i]+pd.DateOffset(hours=9)
        while True:
            #start = start_of_wk.strftime('%H:%M:%S')
            if start_of_wk.strftime('%H:%M:%S') != "17:00:00" and start_of_wk <= end_of_wk:
                sheet.cell(row=cnt, column=2).value = start_of_wk.strftime('%A') + "(" + start_of_wk.strftime('%H:%M:%S') + "-" + (start_of_wk+pd.DateOffset(seconds=10)).strftime('%H:%M:%S') + ")"
                start_of_wk = start_of_wk+pd.DateOffset(seconds=10)
                if start_of_wk.strftime('%H:%M:%S') == "17:00:00":
                    if start_of_wk == end_of_wk:
                        sheet.merge_cells(start_row=wk_cnt, start_column=1, end_row=cnt, end_column=1)
                        sheet.cell(row=wk_cnt, column=1).value = "Week"+str(usr)
                        wk_cnt = cnt+1
                        usr = usr+1
                    start_of_wk = start_of_wk+pd.DateOffset(hours=15)
                cnt = cnt+1
            else:
                break
        i=i+1
    wbook.save(result_file)


def main():
    #df1 = pd.read_excel(excel_file_1, 'Sheet1')
    #
    # print(df1.columns)
    #
    # filt = (df1['Duration'] != 0)
    # filtered_df = df1.loc[filt]
    # print(filtered_df.head(10))
    #
    # # create excel writer
    # #writer = pd.ExcelWriter(excel_file_1)
    # # book = load_workbook(excel_file_1)
    # # writer = pd.ExcelWriter(excel_file_1, engine='openpyxl')
    # # writer.book = book
    #
    # append_df_to_excel(excel_file_1, filtered_df, sheet_name='Final', index=False)
    #
    # # writer.save()
    #
    # df2 = pd.read_excel(excel_file_1, 'Final')
    # print(df2.columns)
    # dict_of_df = df2.to_dict()
    # dict_rfp_values = dict_of_df['Real First Packet'].values()
    # result = []
    # for value in dict_rfp_values:
    #     # print(value)
    #     result.append(datetime.datetime.fromtimestamp(value / 1000.0).strftime('%Y-%m-%d %H:%M:%S'))
    # df2['RFP'] = result
    # df2['Octet/Duration'] = df2['doctets'] / df2['Duration']
    # #
    # filtered_df2 = df2.sort_values("RFP")
    # print(filtered_df2.head(10))
    #
    # append_df_to_excel(excel_file_1, filtered_df2, sheet_name='Final', index=False, startrow=0)
    df2 = pd.read_excel(excel_file_1, 'Final')
    dt_df = pd.to_datetime(df2['RFP'])
    dict_of_df = dt_df.to_dict()
    dict_rfp_values = dict_of_df.values()
    start_weeks = []
    for value in dict_rfp_values:
        #print(value)
        dt = value
        #print(dt.strftime('%A'))
        if dt.strftime('%A') == 'Monday' and start_weeks.__contains__(dt.strftime("%m/%d/%y")) is False:
            start_weeks.append(dt.strftime("%m/%d/%y"))
            #start_week.append(pd.to_datetime(dt.strftime("%m/%d/%y"))+ pd.DateOffset(hours=8))
    print(start_weeks)
    start_week = []
    for dt in start_weeks:
        start_week.append(pd.to_datetime(dt) + pd.DateOffset(hours=8))
    print(start_week)
    end_week = []
    for dt in start_week:
        #date_1 = datetime.datetime.strptime(dt, "%m/%d/%y")
        end_week.append(pd.to_datetime(dt) + pd.DateOffset(days=4))
    print(end_week)
    result_file_data(start_week, end_week)
    #
    # filtered_df2.to_excel(writer, sheet_name='Sheet1', index=False)
    #wb = writer.book
    #wb.active = 1
    #ws = writer.sheets['Final']
    #ws.column_dimensions['K'].width = 20

    #auto_adjust_xlsx_column_width(filtered_df2, writer, sheet_name="Final", margin=0)
    # writer.save()
    # writer.close()


if __name__ == "__main__":
    main()

